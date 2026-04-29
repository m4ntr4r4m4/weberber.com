"""
Sitemap Product Scraper → Excel
Parses a Shopify sitemap XML, fetches product details (name, price, image),
and exports everything to a formatted .xlsx file with embedded thumbnails.

Usage:
    python sitemap_to_excel.py
    python sitemap_to_excel.py --url "https://..." --output products.xlsx
"""

import argparse
import io
import time
import urllib.parse
from pathlib import Path

import requests
from lxml import etree
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# ── Config ─────────────────────────────────────────────────────────────────────
SITEMAP_URL = (
    "https://beldinest.com/sitemap_products_1.xml?from=1221033164857&to=8038586581153"
)
OUTPUT_FILE = "beldinest_products.xlsx"
ROW_HEIGHT_PX = 80          # Thumbnail height in the sheet (pixels)
COL_IMG_WIDTH = 14          # Column width in Excel units
DELAY_BETWEEN_REQUESTS = 0.5  # Seconds between HTTP calls (be polite)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# ── Helpers ────────────────────────────────────────────────────────────────────

def get(url: str, retries: int = 3, **kwargs) -> requests.Response:
    """GET with simple retry + back-off."""
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=HEADERS, timeout=15, **kwargs)
            r.raise_for_status()
            return r
        except requests.HTTPError as exc:
            if exc.response.status_code == 429 and attempt < retries - 1:
                wait = 2 ** (attempt + 2)
                print(f"  Rate-limited, waiting {wait}s …")
                time.sleep(wait)
            else:
                raise
    raise RuntimeError(f"Failed to GET {url}")


def parse_sitemap(url: str) -> list[dict]:
    """
    Parse the sitemap XML.
    Shopify sitemaps include <image:loc>, <image:title> alongside <loc>.
    Returns list of {url, image_url, title}.
    """
    print(f"Fetching sitemap: {url}")
    r = get(url)
    root = etree.fromstring(r.content)

    ns = {
        "sm": "http://www.sitemaps.org/schemas/sitemap/0.9",
        "image": "http://www.google.com/schemas/sitemap-image/1.1",
    }

    products = []
    for url_el in root.findall("sm:url", ns):
        loc = url_el.findtext("sm:loc", namespaces=ns) or ""
        img_el = url_el.find("image:image", ns)
        image_url = img_el.findtext("image:loc", namespaces=ns) if img_el is not None else ""
        title = img_el.findtext("image:title", namespaces=ns) if img_el is not None else ""
        if loc:
            products.append({"url": loc, "image_url": image_url or "", "title": title or ""})

    print(f"Found {len(products)} product URLs in sitemap")
    return products


def shopify_handle(product_url: str) -> str:
    """Extract Shopify product handle from URL."""
    path = urllib.parse.urlparse(product_url).path  # /products/handle
    return path.rstrip("/").split("/")[-1]


def fetch_price(product_url: str) -> str:
    """
    Fetch price via Shopify's JSON endpoint (/products/<handle>.json).
    Returns formatted price string or '—' on failure.
    """
    handle = shopify_handle(product_url)
    base = f"{urllib.parse.urlparse(product_url).scheme}://{urllib.parse.urlparse(product_url).netloc}"
    json_url = f"{base}/products/{handle}.json"
    try:
        data = get(json_url).json()
        variants = data.get("product", {}).get("variants", [])
        if variants:
            price = variants[0].get("price", "")
            # Shopify returns price as string like "29.99"
            return f"${float(price):,.2f}" if price else "—"
    except Exception as exc:
        print(f"    ⚠ Could not fetch price for {handle}: {exc}")
    return "—"


def download_thumbnail(image_url: str, size: int = 200) -> bytes | None:
    """Download image and resize to thumbnail; returns PNG bytes or None."""
    if not image_url:
        return None
    # Request a smaller variant from Shopify CDN if possible
    thumb_url = image_url.split("?")[0] + f"?width={size}"
    try:
        r = get(thumb_url)
        img = PILImage.open(io.BytesIO(r.content)).convert("RGB")
        img.thumbnail((size, size), PILImage.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()
    except Exception as exc:
        print(f"    ⚠ Thumbnail failed: {exc}")
        return None


# ── Excel builder ──────────────────────────────────────────────────────────────

BRAND_COLOR  = "1A3C5E"   # Dark navy header
ACCENT_COLOR = "E8F0F8"   # Light blue row stripe
WHITE        = "FFFFFF"

def build_excel(products: list[dict], output: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # ── Column widths
    ws.column_dimensions["A"].width = COL_IMG_WIDTH   # Image
    ws.column_dimensions["B"].width = 45              # Name
    ws.column_dimensions["C"].width = 14              # Price
    ws.column_dimensions["D"].width = 50              # URL

    # ── Header row
    header_font  = Font(name="Arial", bold=True, color=WHITE, size=11)
    header_fill  = PatternFill("solid", start_color=BRAND_COLOR)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Image", "Product Name", "Price", "URL"]
    ws.row_dimensions[1].height = 28
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # ── Data rows
    row_px_to_excel = 0.75   # approx conversion factor
    thumb_h = int(ROW_HEIGHT_PX / row_px_to_excel)

    data_font   = Font(name="Arial", size=10)
    stripe_fill = PatternFill("solid", start_color=ACCENT_COLOR)
    plain_fill  = PatternFill("solid", start_color=WHITE)
    url_font    = Font(name="Arial", size=9, color="1155CC", underline="single")

    for i, product in enumerate(products):
        excel_row = i + 2
        fill = stripe_fill if i % 2 == 0 else plain_fill
        ws.row_dimensions[excel_row].height = ROW_HEIGHT_PX

        # Col B – Name
        name_cell = ws.cell(row=excel_row, column=2, value=product.get("title") or product.get("name", ""))
        name_cell.font = Font(name="Arial", size=10, bold=True)
        name_cell.alignment = Alignment(vertical="center", wrap_text=True)
        name_cell.fill = fill
        name_cell.border = border

        # Col C – Price
        price_cell = ws.cell(row=excel_row, column=3, value=product.get("price", "—"))
        price_cell.font = Font(name="Arial", size=11, bold=True, color="1A3C5E")
        price_cell.alignment = Alignment(horizontal="center", vertical="center")
        price_cell.fill = fill
        price_cell.border = border

        # Col D – URL
        url_cell = ws.cell(row=excel_row, column=4, value=product.get("url", ""))
        url_cell.font = url_font
        url_cell.alignment = Alignment(vertical="center", wrap_text=True)
        url_cell.fill = fill
        url_cell.border = border

        # Col A – Image (empty fill first)
        img_cell = ws.cell(row=excel_row, column=1, value="")
        img_cell.fill = fill
        img_cell.border = border

        thumb_bytes = product.get("thumb")
        if thumb_bytes:
            try:
                xl_img = XLImage(io.BytesIO(thumb_bytes))
                xl_img.width  = int(ROW_HEIGHT_PX * 1.0)
                xl_img.height = int(ROW_HEIGHT_PX * 1.0)
                cell_addr = f"A{excel_row}"
                ws.add_image(xl_img, cell_addr)
            except Exception as exc:
                print(f"    ⚠ Could not embed image row {excel_row}: {exc}")

        pct = (i + 1) / len(products) * 100
        print(f"  [{pct:5.1f}%] Row {excel_row}: {product.get('title', '')[:60]}")

    # ── Freeze header row
    ws.freeze_panes = "A2"

    # ── Auto-filter
    ws.auto_filter.ref = f"A1:D{len(products) + 1}"

    wb.save(output)
    print(f"\n✅  Saved → {output}  ({len(products)} products)")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Shopify sitemap → Excel")
    parser.add_argument("--url",    default=SITEMAP_URL,  help="Sitemap URL")
    parser.add_argument("--output", default=OUTPUT_FILE,  help="Output .xlsx path")
    parser.add_argument("--no-images", action="store_true", help="Skip image download")
    parser.add_argument("--limit",  type=int, default=0,  help="Max products (0=all)")
    args = parser.parse_args()

    # 1. Parse sitemap
    products = parse_sitemap(args.url)
    if args.limit:
        products = products[: args.limit]

    # 2. Enrich each product
    for idx, p in enumerate(products):
        handle = shopify_handle(p["url"])
        print(f"[{idx+1}/{len(products)}] {handle}")

        # Price
        p["price"] = fetch_price(p["url"])
        time.sleep(DELAY_BETWEEN_REQUESTS)

        # Thumbnail
        if not args.no_images and p["image_url"]:
            p["thumb"] = download_thumbnail(p["image_url"])
            time.sleep(DELAY_BETWEEN_REQUESTS * 0.5)
        else:
            p["thumb"] = None

    # 3. Build Excel
    build_excel(products, args.output)


if __name__ == "__main__":
    main()
